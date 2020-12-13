import openpyxl as xl

# We need to start with a workbook object
# We can create an empty workbook it in Memory or open an existing one in disk

# To create an empty wb in memory with use the Workbook method
# wb = xl.Workbook()

# To open an existing one we use the load_workbook method
wb = xl.load_workbook("transactions.xlsx")

# There's only one sheet called sheet 1
print(wb.sheetnames)

# To create use create_sheet method
# To remove use remove_sheet method

# We access the sheet just as we would a dict
sheet = wb["Sheet1"]

# We can access cell A1 like so
cell = sheet["a1"]

# The cell attr gives us the value in the cell
print(cell.value)

# We can change the value like so:
# cell.value = 1

print(cell.row)
print(cell.column)
print(cell.coordinate)

# We can also access the cell using the cell method of the sheet object
cell = sheet.cell(row=1, column=1)

# We can easly iterate over rows and columns
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


# Access all cells in the A column
column = sheet["a"]
print(column)

# We can work with a range
cells = sheet["a:c"]
print(cells)

# We can work with a range (coordinates)
cells = sheet["a1:c3"]
print(cells)

# Add row at the end of the sheet with the append method
sheet.append([1, 2, 4])

# Save wb as new files
wb.save("transactions2.xlsx")
